import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import tkinter as tk
from tkinter import filedialog, messagebox


def process_workbook(filename):
    try:
        wb = xl.load_workbook(filename)
        sheet = wb.active

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 3)
            value = cell.value

            corrected_price = float(value) * 0.9
            sheet.cell(row, 4).value = corrected_price

        values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, 'e2')

        wb.save(filename)
        return True
    except Exception as e:
        print(f"Eroare: {e}")
        return False


def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        success = process_workbook(file_path)
        if success:
            messagebox.showinfo("Succes")
        else:
            messagebox.showerror("Eroare")



root = tk.Tk()
root.title("Excel Processor")
root.geometry("300x150")

label = tk.Label(root, text="Selectează fișierul:", pady=20)
label.pack()

btn = tk.Button(root, text="Încarcă Fișier", command=select_file, bg="#4CAF50", fg="white", padx=10)
btn.pack()

root.mainloop()